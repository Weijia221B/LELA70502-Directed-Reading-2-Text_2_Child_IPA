import re
import os
from openpyxl import load_workbook, Workbook
from tqdm import tqdm

def clean_text(text):
    """清理text列的内容"""
    if not isinstance(text, str):
        return ""
    
    # 删除原始时间戳（如"967820_969961"）
    text = re.sub(r'\d+_\d+', '', text)
    
    # 删除所有圆括号及括号内的内容
    text = re.sub(r'\([^)]*\)', '', text)
    
    # 删除所有方括号及括号内的内容
    text = re.sub(r'\[[^\]]*\]', '', text)
    
    # 删除 [?] 标记
    text = re.sub(r'\[\?\]', '', text)
    
    # 删除 @l 标记
    text = re.sub(r'@l', '', text)
    
    # 删除 // 标记
    text = re.sub(r'//', '', text)
    
    # 删除特殊字符和数字
    text = re.sub(r'&-|～|=|\+|\(\.\.\.\)|\d+', '', text)
    
    # 清理多余空格
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def clean_utterance(utterance):
    """清理child_utterance列并添加单词边界标记"""
    if not isinstance(utterance, str):
        return ""
    
    # 删除特殊字符
    utterance = re.sub(r'&-|～|=|\+|\(\.\.\.\)', '', utterance)
    
    # 清理多余空格
    utterance = re.sub(r'\s+', ' ', utterance).strip()
    
    # 先用特殊符号标记单词边界
    words = utterance.split()
    if len(words) > 1:
        utterance = "|".join(words)  # 使用 | 作为临时分隔符
    
    # 在音素之间添加空格
    utterance = ' '.join(utterance)
    
    # 将临时分隔符替换为 WORD_BOUNDARY
    utterance = utterance.replace('|', ' WORD_BOUNDARY ')
    
    return utterance

def main():
    # 创建输出目录(如果不存在)
    os.makedirs("output", exist_ok=True)
    
    input_file = "data/summary1_168303.xlsx"
    output_file = "output/processed_data_age_1_2_v2.xlsx"
    
    try:
        # 加载工作簿时避免数据类型转换
        wb = load_workbook(filename=input_file, data_only=True)
        # 选择工作表
        ws = wb["age1_2"]
        
        print(f"成功读取数据，共 {ws.max_row-1} 行")
        
        # 获取列索引
        header_row = next(ws.rows)
        column_indices = {}
        for i, cell in enumerate(header_row):
            column_indices[cell.value] = i
            
        # 创建新的工作簿
        new_wb = Workbook()
        new_ws = new_wb.active
        
        # 写入标题行
        new_ws.append(["Number", "text", "child_utterance"])
        
        # 处理数据行并显示进度条
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in tqdm(rows, desc="处理数据"):
            number = row[column_indices["Number"]]
            word_spoken = row[column_indices["Word_spoken"]]
            child_phon = row[column_indices["Child_Phon"]]
            
            # 应用清理函数
            cleaned_text = clean_text(word_spoken)
            cleaned_utterance = clean_utterance(child_phon)
            
            # 添加到新工作表
            new_ws.append([number, cleaned_text, cleaned_utterance])
        
        # 保存新工作簿，设置不自动调整格式
        new_wb.save(output_file)
        print(f"数据处理完成，结果已保存到 {output_file}")
        
    except Exception as e:
        print(f"处理Excel文件时出错: {e}")
        return

if __name__ == "__main__":
    main() 